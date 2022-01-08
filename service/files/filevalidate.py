import os
from openpyxl import load_workbook

class Filevalidate():
    def file_exists_check (self, filename):
        return os.path.isfile (filename)

    def non_empty_file_check (self,filename):
        return os.path.getsize (filename) > 0

    def file_extension_check (self, filename):
        file_extention=os.path.splitext (filename) [1] [1:]
        return file_extention in ("xlsx","csv","txt","json")
    def check_file_type(self,filename,fileType):
        expectedType=filename.split('.')[-1]
        return expectedType==fileType

#update excel sheet name into data instance        
def update_sheet_name(payload, name):
    data_ins=[]
    for rec in payload['dataInstances'][0]['dataParameters']:
        r=[]
        r=rec.copy()
        if rec['fieldName']=='sheet_name':
            r['userValue']=name
        data_ins.append(r)
    payload['dataInstances'][0]['dataParameters']=data_ins
    return payload['dataInstances'][0]

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def update_excel_data_param(payload,filePath):
    sheetNames = get_sheetnames_xlsx(filePath)
    data_ins=[]
    for i,name in enumerate(sheetNames):
        payload['dataInstances'][0]['id']='dataInstances_'+str(i+1)
        payload['dataInstances'][0]['name']=name
        data_ins.append(update_sheet_name(payload, name).copy())
    return data_ins